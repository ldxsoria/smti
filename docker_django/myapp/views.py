from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.models import User
from django.contrib.auth.forms import  AuthenticationForm #SIGNIN

#MODELOS
from .models import Ticket, EstadosTicket, Registro, Area, Asunto, Reporte
from inventory.models import Activo

#LISTVIEW REQUIREMENTS
from django.views.generic import ListView
from django.core.paginator import Paginator #PAGINATION

#SEARCH REQUIEREMENTS
from django.db.models import Q

#PROJECTS ROUTES
from django.contrib.auth import login, logout, authenticate #para crear cookie de inicio de sesion
from django.contrib.auth.decorators import login_required #MAIN

#MY REQUIREMENTS
from .forms import TicketForm, RegistroForm, ReporteForm

#IMPORT CSV REQUIREMENTS
import csv, io
from import_export import resources
from django.contrib.auth.hashers import make_password #USER > PASSWORD

#REQUISITOS PARA EL CORREO
from django.template.loader import get_template
from django.core.mail import EmailMultiAlternatives
from django.conf import settings
from django.contrib.sites.shortcuts import get_current_site #para obtener el dominio actual
import threading
from decouple import config #CORREOS DE CC COMO VARIABLES DE ENTORNO

#REPORTE PDF
#from django.http import HttpResponse
from reportlab.pdfgen import canvas

#REPORTE EXCEL
#from django.http import HttpResponse
from datetime import datetime #PARA LA HORA
from openpyxl import Workbook

#ENVIAR MENSAJES DE UNA VISTA A OTRA
from django.contrib import messages

#TELEGRAM
# import requests as telegram

#FUNCIONES_GENERALES##################################################################################
def create_mail(user, cc_mails, subject, template_path, context):

    template = get_template(template_path)
    content = template.render(context)

    mail = EmailMultiAlternatives(
        subject=subject,
        body=content,
        from_email=settings.EMAIL_HOST_USER,
        to=[
            user.email,
        ],
        cc= cc_mails
    )

    mail.attach_alternative(content, 'text/html')
    mail.send(fail_silently=False)
    #return mail


##########################################################################################################

# Create your views here.
def signin(request):
    #Si esta autenticado la pagina principal debe de ser main
    if request.user.is_authenticated:
        activos = Activo.objects.filter(responsable=request.user).exists()
        context = {
            'activos' : activos
        }
        return render(request, 'main.html', context)
    else:
        if request.method == 'GET':
            return render(request, 'signin.html',{
            'form': AuthenticationForm
        })
        else:
            user = authenticate(request, username=request.POST['username'], password=request.POST['password'])

            if user is None:
                return render(request, 'signin.html',{
                'form': AuthenticationForm,
                'error': 'Username or password is incorrect'
            })
            else:
                login(request, user)
                return redirect('main/')

@login_required
def signout(request):
    logout(request)
    return redirect('/')

@login_required
def main(request):
    activos = Activo.objects.filter(responsable=request.user).exists()
    context = {
        'activos' : activos,
    }
    return render(request, 'main.html', context)

#TICKET###############################################################################################

@login_required      
def tickets(request):
    if request.user.is_staff:
        #https://stackoverflow.com/questions/7590692/django-get-unique-object-list-from-queryset
        tickets = Ticket.objects.all().exclude(completado=True).order_by('-id')
        areas = Area.objects.all()
        return render(request, 'tickets/tickets.html', {
            'tickets': tickets,
            'areas':areas,
            'title': 'Tickets nuevos'
        })
    else:
        tickets = Ticket.objects.filter(solicitante_id=request.user.id).exclude(completado=True)
        areas = Area.objects.all()
        return render(request, 'tickets/tickets.html', {
            'tickets': tickets,
            'areas':areas,
            'title': 'Mis tickets pendientes',
        })

@login_required
def completed_tickets(request):
    if request.user.is_staff:
        tickets = Ticket.objects.filter(completado=True).order_by('-id')
        for ticket in tickets:
            ticket = Ticket.objects.get(id=ticket.id)
            registro =  ticket.registro.order_by('-id')[:1].values()
            read = Registro.objects.filter(ticket__id=ticket.id).order_by('-id')[:1]
        
        return render(request, 'tickets/completed_tickets.html', {
            'tickets': tickets,
            'title': 'Tickets completados',
        })
    else:
        tickets = Ticket.objects.filter(solicitante_id=request.user.id, completado=True).order_by('-id')
        return render(request, 'tickets/completed_tickets.html', {
            'tickets': tickets,
            'title': 'Mis tickets completados'
        })


class SearchCreatedTickets(ListView):
    paginate_by = 10
    model = Ticket
    template_name = 'tickets/search_result_tickets.html'

    def get_queryset(self):  # new
        query = self.request.GET.get("q")
        if query is None:
            return Ticket.objects.all()
        else:
            print(query)
            object_list = Ticket.objects.filter(
                #Q(razon__startswith=query) | Q(razon__icontains=query)
                #Q(id__icontains=query) | Q(asunto__icontains=query)
                Q(id__icontains=query)
            )
            print(object_list)
            return object_list

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Tickets creados'
        #context['lugar'] = Area.objects.all().prefetch_related('ticket')
        context['lugar'] = Area.objects.all()
        return context
    
class SearchCompletedTickets(ListView):
    paginate_by = 10
    model = Ticket
    template_name = 'tickets/search_result_tickets.html'

    def get_queryset(self):  # new
        query = self.request.GET.get("q")
        if query is None:
            if self.request.user.is_staff:
                return Ticket.objects.filter(completado=True)
            else:
                return Ticket.objects.filter(completado=True, solicitante_id=self.request.user)
        else:
            print(query)
            object_list = Ticket.objects.filter(
                Q(id__icontains=query)
            )
            print(object_list)
            return object_list
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Tickets completados'
        context['lugar'] = Area.objects.all()
        return context

@login_required
def create_ticket(request):
    if request.method == 'GET':
        areas = Area.objects.all()
        asuntos = Asunto.objects.all()
        return render(request, "tickets/create_ticket.html", {
            'form': TicketForm,
            'areas': areas,
            'asuntos': asuntos
        })
    else:
        try:
            #VERIFICAR SI LUGAR & ASUNTO SE SELECCIONARON ANTES DE HACER POST
            if  ('lugar' not in request.POST)  or  ('asunto' not in request.POST):
                areas = Area.objects.all()
                asuntos = Asunto.objects.all()              
                context = {
                    'areas': areas,
                    'asuntos': asuntos,
                    'type' : 'danger',
                    'msg' : '¡Seleccione un asunto y el lugar del incidente!'
                }

                return render(request, "tickets/create_ticket.html",context)
            
            else:
                asunto = Asunto.objects.filter(id=request.POST['asunto'])
                new_ticket = Ticket(
                    asunto=Asunto.objects.get(id=request.POST['asunto']),
                    descripcion=request.POST['descripcion'],
                    solicitante=request.user,
                    )
                new_ticket.save()
                #ASIGNAR TICKET AL LUGAR SELECCIONADO
                area_selected = Area.objects.get(cod_area=request.POST['lugar'])
                #area_selected.add(ticket=new_ticket)
                area_selected.ticket.add(new_ticket)
                area_selected.save()
                #
                new_ticket.save()
                new_registro = Registro(responsable=request.user, estado=EstadosTicket(estado=1), comment_estado='REGISTRO AUTOMATICO')
                new_registro.save()
                new_ticket.registro.add(new_registro)
                new_ticket.save()
                #FILTAR LUGAR PARA ENVIARLO POR CORREO
                query_lugar_ticket = Area.objects.filter(ticket = new_ticket)
                lugar_ticket = query_lugar_ticket.values()[0]['descripcion']
                #------------
                #ENVIAR CORREO
                cc_admins = config('CC_TICKETS') #LOS OBTENGO DEL .env
                cc_mails = cc_admins.split(",")# COMO SIEMPRE SERA STR, LO TRANSFORMO EN UNA LISTA
                subject= f'Ticket #{new_ticket.id}'
                template_path = 'tickets/correo_new_ticket.html'
                dominio = get_current_site(request).domain
                context = {
                    'user' : request.user,
                    'ticket' : new_ticket,
                    'lugar': lugar_ticket,
                    'dominio' : dominio
                }
                #ENVIO NORMAL
                #new_ticket_mail = create_mail(request.user, cc_mails, subject, template_path, context)
                #new_ticket_mail.send(fail_silently=False)
                
                thread = threading.Thread(
                    target= create_mail,
                    args=(request.user, cc_mails, subject, template_path, context)
                )
                thread.start()
                #------------
                #NOTIFICAR POR TELEGRAM
                # telegram.post('https://api.telegram.org/bot6010368859:AAH3LbF3Z8SlIDWKGVCd4ZxrloSaE-1CQUQ/sendMessage', data = {'chat_id' : '@mtilogs', 'text' : 'Nuevo ticket registrado'})
                #------------
                context = {
                        'type' : 'success',
                        'alert' : '¡El ticket fue registrado con exito!'
                    }
                return render(request, 'main.html',context)
                #return redirect(f'{new_ticket.id}/progress')
        except ValueError as e:
            return render(request, 'tickets/create_ticket.html', {
                'form': TicketForm,
                'error': f'Please provide valida data > {e}'
            })

@login_required
def historial_ticket(request, ticket_id):
    if request.method == 'GET':
        registros = Registro.objects.filter(ticket__id=ticket_id).order_by('-hora_estado', 'fecha_estado')
        ticket = get_object_or_404(Ticket, pk=ticket_id)

        return render(request, 'tickets/historial_ticket.html',{
            'registros':registros,
            'ticket': ticket,
        })

@login_required
def progress_ticket(request, ticket_id):
    if request.user.is_staff:
        if request.method == 'GET':
            registros = Registro.objects.filter(ticket__id=ticket_id).order_by('-hora_estado', 'fecha_estado')
            ticket = get_object_or_404(Ticket, pk=ticket_id)
            #form = TicketForm(instance=ticket)
            formAddRegistro = RegistroForm(instance=ticket)
            estados = EstadosTicket.objects.all()
            areas = Area.objects.all()
            #formReporte = ReporteForm()

            area_actual = Area.objects.filter(ticket=ticket_id)
            print(area_actual)

            reporte_existe = Reporte.objects.filter(ticket=ticket_id).exists()
            reporte_actual = Reporte.objects.filter(ticket=ticket_id).values()[0] if reporte_existe == True else None

            return render(request, 'tickets/progress_ticket.html',{
                'registros':registros,
                'ticket': ticket,
                #'form': form,
                'formAddRegistro' : formAddRegistro,
                'estados': estados,
                'areas': areas,
                'area_actual':area_actual,
                #'formReporte': formReporte,
                'reporte': reporte_actual,
            })
        else:
            try:
                registros = Registro.objects.filter(ticket__id=ticket_id).order_by('-hora_estado', 'fecha_estado')
                ticket = get_object_or_404(Ticket, pk=ticket_id)
                #CREANDO UN REPORTE
                new_reporte = Reporte (
                    contexto = request.POST['contexto'],
                    diagnostico = request.POST['diagnostico'],
                    recomendacion = request.POST['recomendacion'],
                    #ticket = Ticket.objects.filter(id=ticket_id),
                    ticket = ticket,
                    created_by = request.user,
                )
                print(new_reporte)
                new_reporte.save()
                return redirect('progress_ticket', ticket_id)
                #form = TicketForm(request.POST or None, instance=ticket)
                #if form.is_valid():
                #    form.save()                
                #fromReporte = ReporteForm(request.POST or None, instance=ticket)
                # if fromReporte.is_valid():
                #     fromReporte.save()
                #     return redirect('progress_ticket', ticket_id)
                # else:
                #     print('ERROR'*100)
            except ValueError:
                return HttpResponse('No funciono ERROR')
    else:
        return HttpResponse('Error #0001 de permiso, comuniquese con el Administrador del sistema')

@login_required
def add_registro_ticket(request, ticket_id): 
    if request.method == 'GET':
        return HttpResponse('ERROR add_registro_ticket')
    else:
        try:
            print('*'*50)
            print(request.POST['estado'])
            #print(request.POST['comentario'])
            #------------
            ticket = Ticket.objects.get(id=ticket_id)
            #MODIFICAR A COMPLETADO SI EL ESTADO ES
            if request.POST['estado'] == '6':
                #HORA Y FECHA DEL CIERRE
                ticket.fecha_cierre = datetime.now().date()
                ticket.hora_cierre = datetime.now().time()
                #
                ticket.completado = True
            new_registro = Registro(responsable=request.user, estado=EstadosTicket(estado=request.POST['estado']), comment_estado=request.POST['comentario'])
            new_registro.save()
            ticket.registro.add(new_registro)
            ticket.save()
            #------------
            return redirect('progress_ticket', ticket_id)
        except ValueError as e:
            return render(request, 'tickets/create_ticket.html', {
                'form': TicketForm,
                'error': f'Please provide valida data > {e}'
            })

@login_required
def add_ticket_to_area(request, ticket_id):
    if request.user.is_staff:
        if request.method == 'POST':
            #OBTENGO EL ID DE LA LA PETICION
            ticket = Ticket.objects.get(id=ticket_id)
            #OBTENGO EL AREA SELECIONA DE FORM POR EL POST
            area = Area.objects.get(cod_area=request.POST['area'])
            #ASIGNO EL TICKET A AREA
            area.ticket.add(ticket)
            area.save()
    
            return redirect('progress_ticket', ticket_id)
        
@login_required
def add_reporte_ticket(request, ticket_id):
    if request.user.is_staff:
        if request.method == 'POST':
            #CREANDO UN REPORTE
            ticket = get_object_or_404(Ticket, pk=ticket_id)
            reporte = Reporte.objects.filter(ticket_id=ticket_id).exists()
            if reporte == False:
                new_reporte = Reporte (
                    contexto = request.POST['contexto'],
                    diagnostico = request.POST['diagnostico'],
                    recomendacion = request.POST['recomendacion'],
                    #ticket = Ticket.objects.filter(id=ticket_id),
                    ticket = ticket,
                    created_by = request.user,
                )
                new_reporte.save()
            else:
                #ACTUALIZO EL REPORTE
                update_reporte = get_object_or_404(Reporte, ticket_id=ticket_id)
                contexto = request.POST['contexto']
                diagnostico = request.POST['diagnostico']
                recomendacion = request.POST['recomendacion']

                # Actualiza los campos del objeto Reporte
                update_reporte.contexto = contexto
                update_reporte.diagnostico = diagnostico
                update_reporte.recomendacion = recomendacion
                update_reporte.save()
                return redirect('progress_ticket', ticket_id)
            
        return redirect('progress_ticket', ticket_id)
        
@login_required
def deactivate_ticket(request, ticket_id):
    if request.method == 'POST':
        ticket = Ticket.objects.get(id=ticket_id)
        ticket.completado = True
        new_registro = Registro(responsable=request.user, estado=EstadosTicket(estado=5))#5 CANCELADO POR EL USUARIO
        new_registro.save()
        ticket.registro.add(new_registro)
        ticket.save()

        return redirect('progress_ticket', ticket_id)

@login_required
def delete_registro_to_ticket(request, ticket_id, registro_id):
    if request.user.is_staff:
        if request.method == 'POST':
            # #OBTENGO EL ID DE LA LA PETICION
            # ticket = Ticket.objects.get(id=ticket_id)
            # #OBTENGO EL AREA SELECIONA DEL FORM
            # area = Area.objects.get(cod_area=cod_area)
            # #ASIGNO EL TICKET A AREA
            # area.ticket.remove(ticket)
            # area.save()
            registro = Registro.objects.get(id=registro_id)
            registro.delete()
    
            return redirect('progress_ticket', ticket_id)
        
@login_required
def delete_ticket_to_area(request, ticket_id, cod_area):
    if request.user.is_staff:
        if request.method == 'POST':
            #OBTENGO EL ID DE LA LA PETICION
            ticket = Ticket.objects.get(id=ticket_id)
            #OBTENGO EL AREA SELECIONA DEL FORM
            area = Area.objects.get(cod_area=cod_area)
            #ASIGNO EL TICKET A AREA
            area.ticket.remove(ticket)
            area.save()
    
            return redirect('progress_ticket', ticket_id)

def dashboard_tickets(request):

    count_tickets = Ticket.objects.count()
    count_tickets_completed = Ticket.objects.filter(completado=True).count()
    count_tickets_uncompleted = Ticket.objects.filter(completado=False).count()

    # cs_no = Students.objects.filter(course='Computer Science').count()
    # cs_no = int(cs_no)
    # print('Number of Computer Science Students Are',cs_no)

    # asuntos_list = Asunto.objects.values_list('id','desc')
    # ticket_list = Ticket.objects.filter(asunto=23).count()

    asuntos = Asunto.objects.values_list('id','desc')

    asuntos_title_list = []
    asuntos_value_list = []


    for a in asuntos :
        total_asunto_x = Ticket.objects.filter(asunto=a[0]).count()
        if total_asunto_x > 0:
            asuntos_title_list.append(a[1])
            asuntos_value_list.append(total_asunto_x)



    content = {
                'count_tickets' : count_tickets,
                'count_tickets_completed' : count_tickets_completed,
                'count_tickets_uncompleted' : count_tickets_uncompleted,
                ########################################################
                'asuntos_title_list' : asuntos_title_list,
                'asuntos_value_list' : asuntos_value_list,

    }

    return render(request, 'tickets/dashboard_tickets.html',content)

#EXPORT VIEWS##########################################################################################

@login_required       
def export_csv(request):
    if request.user.is_staff:
        tickets_resource = resources.modelresource_factory(model=Ticket)()
        dataset = tickets_resource.export()
        response = HttpResponse(dataset.csv, content_type='text/csv')
        response['Content-Disposition'] = 'atachment; filename="tickets_export.csv"'
        return response

#IMPORT VIEWS##########################################################################################

@login_required    
def import_csv(request):

    if request.user.is_staff:
        #libreria import_export
        import tablib

        with open("example.csv", "r") as csv_file:
            usuarios_resource = resources.modelresource_factory(model=User)()
            dataset = tablib.Dataset(headers=[field.name for field in User._meta.fields]).load(csv_file)
            result = usuarios_resource.import_data(dataset, dry_run=True)
        if not result.has_errors():
            usuarios_resource.import_data(dataset, dry_run=False)
        return HttpResponse(
            "Users successfully imported"
        )

@login_required
def auto_import(request, model):
    if request.user.is_staff:
        template = 'general/import.html'
        context = {
            'title': model
        }
        if model == 'areas' or model == 'users' or model == 'estados-ticket' or model == 'asunto':
            if request.method == 'GET':
                return render(request, template, context)
            try:
                csv_file = request.FILES['file']
                data_set = csv_file.read().decode('UTF-8')
                io_string = io.StringIO(data_set)
                next(io_string)

                if model == 'areas':
                    for column in csv.reader(io_string, delimiter=',', quotechar='|'):
                        created = Area.objects.update_or_create(
                            cod_area=column[0],
                            descripcion=column[1],
                            siglas=column[2],
                        )

                elif model == 'users':
                    for column in csv.reader(io_string, delimiter=',', quotechar='|'):
                        created = User.objects.update_or_create(
                            username=column[0],
                            first_name=column[1],
                            last_name=column[2],
                            email=column[3],
                            password=make_password(column[4])
                        )
                elif model == 'estados-ticket':
                    for column in csv.reader(io_string, delimiter=',', quotechar='|'):
                        created = EstadosTicket.objects.update_or_create(
                            estado=column[0],
                            desc=column[1],
                        )
                elif model == 'asunto':
                    for column in csv.reader(io_string, delimiter=',', quotechar='|'):
                        created = Asunto.objects.update_or_create(
                            desc=column[0],
                            tipo=column[1],
                        )                 

                else:
                    return redirect('main')

                context = {
                    'type' : 'success',
                    'alert' : '¡El CSV fue cargardo con exito!',
                    'title': model
                }
                return render(request, template, context)
            except:
                context = {
                'type' : 'danger',
                'alert' : 'Selecciona un .CSV',
                'title': model
                }
                return render(request, template, context)
        else:
            return redirect('main')
            
#-----------------------------------
#REPORTES     ##########################################################################################
@login_required
def report_all_tickets_pdf(request):
    # Creamos un objeto HttpResponse con el tipo de contenido "application/pdf"
    response = HttpResponse(content_type='application/pdf')
    # Indicamos que el archivo se descargará como "reporte.pdf"
    response['Content-Disposition'] = 'attachment; filename="reporte.pdf"'

    # Creamos el objeto PDF utilizando reportlab
    p = canvas.Canvas(response)

    # Agregamos contenido al PDF
    p.drawString(20, 100, "¡Hola, este es un reporte PDF generado desde Django!")

    # Finalizamos el objeto PDF
    p.showPage()
    p.save()

    # Devolvemos la respuesta HTTP con el contenido del PDF
    return response

@login_required
def report_all_tickets_xls(request):
   # Creamos un objeto HttpResponse con el tipo de contenido "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # Indicamos que el archivo se descargará como "reporte.xlsx"
    fecha = datetime.now().date()
    response['Content-Disposition'] = f'attachment; filename="MTI - Reporte #1 - {fecha}.xlsx"'

    # Creamos el libro de trabajo (workbook) de Excel
    workbook = Workbook()
    # Seleccionamos la hoja activa
    hoja = workbook.active

    # Obtenemos todas las áreas y los tickets relacionados
    #areas = Area.objects.all()
    tickets = Ticket.objects.select_related('area','asunto', 'registro').values_list(
        'id', #1
        'area__descripcion', #2
        'fecha_solicitud', #3
        'hora_solicitud', #4
        'fecha_cierre', #5
        'hora_cierre', #6
        'asunto__desc', #7
        'asunto__tipo', #8
        'solicitante_id__username', #9
        'descripcion', #10
        'completado', #11
        # 'registro', #12
        # 'registro__estado__desc', #13
        ).order_by('id')

    #tickets = Ticket.objects.all()
    #areas = Area.objects.select_related('ticket').values_list('area__siglas','id')

    # Agregamos los encabezados al reporte
    hoja['A1'] = "ticket" #1
    hoja['B1'] = "area" #2
    hoja['C1'] = 'fecha_solicitud' #3
    hoja['D1'] = 'hora_solictud' #4
    hoja['E1'] = 'fecha_cierre' #5
    hoja['F1'] = 'hora_cierre' #6
    hoja['G1'] = 'asunto_desc' #7   
    hoja['H1'] = 'asunto_tipo' #8
    hoja['I1'] = 'solicitante_id' #9
    hoja['J1'] = 'descripcion_ticket' #10
    hoja['K1'] = 'completado' #11
    # hoja['L1'] = 'registro_id' #12
    # hoja['M1'] = 'registro_id_estado' #13

    # Agregamos los datos al reporte
    fila = 2
    for t in tickets:
        hoja.cell(row=fila, column=1).value = t[0]
        hoja.cell(row=fila, column=2).value = t[1]
        hoja.cell(row=fila, column=3).value = t[2]
        hoja.cell(row=fila, column=4).value = t[3]
        hoja.cell(row=fila, column=5).value = t[4]
        hoja.cell(row=fila, column=6).value = t[5]
        hoja.cell(row=fila, column=7).value = t[6]
        hoja.cell(row=fila, column=8).value = t[7]
        hoja.cell(row=fila, column=9).value = t[8]
        hoja.cell(row=fila, column=10).value = t[9]
        hoja.cell(row=fila, column=11).value = t[10]
        # hoja.cell(row=fila, column=12).value = t[11]
        # hoja.cell(row=fila, column=13).value = t[12]
        fila += 1

    # Guardamos el libro de trabajo
    workbook.save(response)

    # Devolvemos la respuesta HTTP con el contenido del archivo Excel
    return response